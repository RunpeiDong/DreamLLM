import os
import re
import json
import argparse
from openpyxl import Workbook


def pope_eval(ans_file, label_file):
    answers = ans_file
    label_list = label_file

    key_list = list(answers.keys())
    for key in key_list:
        text = answers[key]

        # Only keep the first sentence
        if text.find('.') != -1:
            text = text.split('.')[0]

        text = text.replace(',', '')
        words = text.split(' ')
        if 'No' in words or 'not' in words or 'no' in words:
            answers[key] = 'no'
        else:
            answers[key] = 'yes'

    for key in key_list:
        if label_list[key] == 'no':
            label_list[key] = 0
        else:
            label_list[key] = 1

    pred_list = {}
    # for answer in answers:
    for key in key_list:
        if answers[key] == 'no':
            pred_list[key] = 0
        else:
            pred_list[key] = 1

    pos = 1
    neg = 0
    pred_list_list = [v for k, v in pred_list.items()]
    yes_ratio = pred_list_list.count(1) / len(pred_list)

    TP, TN, FP, FN = 0, 0, 0, 0
    for key in key_list:
        pred, label = pred_list[key], label_list[key]
    # for pred, label in zip(pred_list, label_list):
        if pred == pos and label == pos:
            TP += 1
        elif pred == pos and label == neg:
            FP += 1
        elif pred == neg and label == neg:
            TN += 1
        elif pred == neg and label == pos:
            FN += 1

    print('TP\tFP\tTN\tFN\t')
    print('{}\t{}\t{}\t{}'.format(TP, FP, TN, FN))

    precision = float(TP) / float(TP + FP)
    recall = float(TP) / float(TP + FN)
    f1 = 2*precision*recall / (precision + recall)
    acc = (TP + TN) / (TP + TN + FP + FN)
    print('Accuracy: {}'.format(acc))
    print('Precision: {}'.format(precision))
    print('Recall: {}'.format(recall))
    print('F1 score: {}'.format(f1))
    print('Yes ratio: {}'.format(yes_ratio))
    return acc, precision, recall, f1, yes_ratio



task_mapping = {"TextVQA": "vqa",
                "VizWizVQA": "vqa",
                "OKVQA": "vqa",
                "VQAv2": "vqa",
                "mmvet": "vqa",
                "DocVQA": "ANLS",
                "InfographicVQA": "ANLS",
                "ST-VQA": "ANLS",
                "TallyQA": "EM",
                "OCR-VQA": "EM",
                "AI2D": "EM",
                "ChartQA": "RA",
                "COCO-Captions": "caption",
                "NoCaps": "caption",
                "TextCaps": "caption",
                "VizWiz-Cap": "caption",
                "Screen2Words": "caption",
                "Image-Paragraph": "caption",
                "OCR": "OCR",
                "POPE_random": "POPE",
                "POPE_popular": "POPE",
                "POPE_adversarial": "POPE",
                "mmbench": "mm"
                }

OCR_prompt = "what is written in the image?"
metric_mapping = {"TextVQA": "accuracy",
                "VizWizVQA": "accuracy",
                "OKVQA": "accuracy",
                "VQAv2": "accuracy",
                "DocVQA": "ANLS",
                "InfographicVQA": "ANLS",
                "ST-VQA": "ANLS",
                "ST-TallyQA": "EM",
                "OCR-VQA": "EM",
                "AI2D": "EM",
                "ChartQA": "RA",
                "COCO-Captions": "CIDEr",
                "Image-Paragraph": "CIDEr",
                "NoCaps": "CIDEr",
                "TextCaps": "CIDEr",
                "VizWiz-Cap": "CIDEr",
                "Screen2Words": "CIDEr",
                "OCR": "accuracy",
                "POPE_random": "accuracy",
                "POPE_popular": "accuracy", 
                "POPE_adversarial": "accuracy",
                "mmbench": "accuracy",
                "mmvet": "accuracy",
                }

class CustomAction(argparse.Action):
        def __call__(self, parser, namespace, values, option_string=None):
            setattr(namespace, self.dest, " ".join(values))

def get_parser():
    
    parser = argparse.ArgumentParser()
    parser.add_argument("--model_name", type=str, default="facebook/opt-350m")
    parser.add_argument("--gtfile_path", type=str, required=True)
    parser.add_argument("--image_path", type=str, required=True)
    parser.add_argument("--out_path", type=str, required=True)
    parser.add_argument("--num-chunks", type=int, default=1)
    parser.add_argument("--chunk-idx", type=int, default=0)
    parser.add_argument("--datatype", type=str, required=True)
    parser.add_argument("--img_aug", type=str, default="center_crop") # center_crop, square_resize, padding_square_resize
    parser.add_argument("--beamsearch", type=str, default="True")
    parser.add_argument("--clip", type=str, default="False")
    parser.add_argument("--evaltype", type=str, default="all")
    parser.add_argument("--task_type", type=str, default="vqa")
    parser.add_argument("--prompt", action=CustomAction, type=str, nargs='+', default=None)
    parser.add_argument('--post_prompt', action=CustomAction, type=str, nargs='+', default=None)
    parser.add_argument('--system_prompt', action=CustomAction, type=str, nargs='+', default=None)
    
    return parser


def merge_outputs(out_path):
    filename = out_path + "/results_final.json"
    files = os.listdir(out_path)

    alist = []
    for file in files:
        if 'final' in file:
            continue
        alist += json.load(open(os.path.join(out_path, file), encoding='utf-8'))

    with open(filename, 'w', encoding="utf-8") as f:
        f.write(json.dumps(alist))


def clear_outputs(out_path):
    if os.path.exists(out_path):
        files = os.listdir(out_path)
        for file in files:
            if 'final' in file:
                continue
            os.remove(os.path.join(out_path, file))


def has_word(sentence, word):
    pattern = r"\b" + re.escape(word) + r"\b"
    match = re.search(pattern, sentence)
    if match:
        return True
    else:
        return False
    
def remove_special_chars(s):
    pattern = r"[^a-zA-Z0-9\s]"
    s = re.sub(pattern, "", s)
    return s

def test_ocr(gt_qa, pre_qa):
    correct = 0
    num = 0
    image_ids = list(gt_qa.keys())
    for image_id in image_ids:
        gt_answers = gt_qa[image_id]
        answer = pre_qa[image_id]
        gt_answers = remove_special_chars(gt_answers).lower()
        answer = remove_special_chars(answer).lower()
        if has_word(answer, gt_answers):
            correct+= 1
        num += 1
    return correct, num


def to_xlsx(json_path, xlsx_path):
    wb = Workbook()
    ws = wb.active

    res = json.load(open(json_path))

    # biaotou
    ws['A1'] = 'question'
    ws['B1'] = 'answer'
    ws['C1'] = 'A'
    ws['D1'] = 'B'
    ws['E1'] = 'C'
    ws['F1'] = 'D'
    ws['G1'] = 'prediction'
    ws['H1'] = 'category'
    ws['I1'] = 'l2-category'
    ws['J1'] = 'index'

    for ann in res:
        row = [ann["question"], ann["answer"], ann["A"], ann["B"], ann["C"], ann["D"], 
            ann["prediction"], ann["category"], ann["l2-category"], ann["index"]]
        ws.append(row)  

    wb.save(xlsx_path)


def process_mmvet(gtfile_path, out_path):
    mm_anns = json.load(open(gtfile_path))
    mm_dict = {}
    for mm_ann in mm_anns:
        mm_dict[mm_ann["question_id"]] = mm_ann['key']

    pred_anns = json.load(open(out_path + "/results_final.json"))
    res = {}
    for pred_ann in pred_anns:
        res[mm_dict[str(pred_ann['question_id'])]] = pred_ann['answer']

    keys = json.load(open("./data/MM-VET/llava_llama2_13b_chat.json"))
    sub_res = {}
    for key in keys.keys():
        sub_res[key] = res[key]

    with open(out_path + "/results_final.json", "w") as f:
        json.dump(res, f, indent=2)